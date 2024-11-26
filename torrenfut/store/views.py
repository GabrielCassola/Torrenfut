from django.http import HttpResponse, Http404
from django.template import loader
from .models import Camiseta, CamisetaTamanho, HistoricoEstoque, Time, Liga, Marca
import json
from django.shortcuts import render, get_object_or_404
from collections import defaultdict
from .models import Camiseta, CamisetaTamanho, TipoProduto
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse

def obter_opcoes_filtro():
    # Obter opções únicas para filtro
    cores_disponiveis = Camiseta.objects.values_list('cor_principal', flat=True).distinct().order_by('cor_principal')
    marcas_disponiveis = Marca.objects.all().order_by('nome')
    times_disponiveis = Time.objects.all().order_by('nome')
    temporadas_disponiveis = Camiseta.objects.values_list('temporada', flat=True).distinct()
    tipos_produto_disponiveis = TipoProduto.objects.all()
    ligas_disponiveis = Liga.objects.all().order_by('nome')

    
    return {
        'cores_disponiveis': cores_disponiveis,
        'marcas_disponiveis': marcas_disponiveis,
        'times_disponiveis': times_disponiveis,
        'temporadas_disponiveis': temporadas_disponiveis,
        'tipos_produto_disponiveis': tipos_produto_disponiveis,
        'ligas_disponiveis': ligas_disponiveis,

    }


def store(request):
    camisetas = Camiseta.objects.all()
    template = loader.get_template('home.html')
    opcoes_filtro = obter_opcoes_filtro()
    context = {
        'camisetas': camisetas,
        **opcoes_filtro,
    }
    return HttpResponse(template.render(context, request))


def produto(request, time_nome, estilo, temporada):
    try:
        # Buscar o time pelo nome, que agora é uma chave estrangeira
        time = get_object_or_404(Time, nome=time_nome)
        
        # Usando get() para buscar a camiseta que corresponde ao time, estilo e temporada
        camiseta = Camiseta.objects.get(time=time, estilo=estilo, temporada=temporada)
        tamanhos = CamisetaTamanho.objects.filter(camiseta=camiseta)
    except Camiseta.DoesNotExist:
        raise Http404("Camiseta não encontrada.")
    
    # Carregar o template
    template = loader.get_template('produto.html')
    
    # Obter as opções de filtro
    opcoes_filtro = obter_opcoes_filtro()
    
    # Contexto com a camiseta, tamanhos e as opções de filtro
    context = {
        'camiseta': camiseta, 
        'tamanhos': tamanhos,
        **opcoes_filtro,
    }
    
    # Retornar a resposta com o contexto
    return HttpResponse(template.render(context, request))

def grafico_estoque(request, produto_id):

    # Obtém a camiseta específica usando o produto_id
    camiseta = get_object_or_404(Camiseta, id=produto_id)

     # Obtendo valores para a página html
    time = camiseta.time
    cor_principal = camiseta.cor_principal
    marca = camiseta.marca
    temporada = camiseta.temporada
    estilo = camiseta.estilo

    # Obtém todos os dados do histórico de estoque
    historico = HistoricoEstoque.objects.all().filter(camiseta=camiseta).order_by('data_alteracao')

    # Organizar os dados por tamanho e data de alteração
    dados_grafico = defaultdict(list)
    for entrada in historico:
        dados_grafico[entrada.produto.tamanho].append([
            entrada.data_alteracao.strftime('%Y-%m-%d'),
            entrada.estoque_novo,
        ])

    # Converter para formato JSON adequado para Google Charts
    dados_formatados = json.dumps({
        tamanho: [["Data", "Estoque Novo"]] + entradas
        for tamanho, entradas in dados_grafico.items()
    })

    context = {
        'dados_grafico': dados_formatados,
        'time': time,
        'estilo': estilo,
        'cor_principal': cor_principal,
        'marca': marca,
        'temporada': temporada,
    }
    return render(request, 'grafico_estoque.html', context)


def filtrar_camisetas(request):
    # Obter filtros da requisição
    filtros_aplicados = {
        'cor': request.GET.get('cor'),
        'marca': request.GET.get('marca'),
        'time': request.GET.get('time'),
        'temporada': request.GET.get('temporada'),
        'tipo_produto': request.GET.get('tipo_produto'),
        'liga': request.GET.get('liga'),
    }

    # Remover filtros que não estão preenchidos
    filtros_aplicados = {chave: valor for chave, valor in filtros_aplicados.items() if valor}

    # Buscar todas as camisetas e aplicar filtros
    camisetas = Camiseta.objects.all()
    if 'cor' in filtros_aplicados:
        camisetas = camisetas.filter(cor_principal=filtros_aplicados['cor'])
    if 'marca' in filtros_aplicados:
        camisetas = camisetas.filter(marca=filtros_aplicados['marca'])
    if 'time' in filtros_aplicados:
        camisetas = camisetas.filter(time=filtros_aplicados['time'])
    if 'temporada' in filtros_aplicados:
        camisetas = camisetas.filter(temporada=filtros_aplicados['temporada'])
    if 'tipo_produto' in filtros_aplicados:
        camisetas = camisetas.filter(tipo_produto__id=filtros_aplicados['tipo_produto'])
    if 'liga' in filtros_aplicados:
        camisetas = camisetas.filter(time__liga=filtros_aplicados['liga'])

    # Obter opções únicas para filtro
    cores_disponiveis = Camiseta.objects.values_list('cor_principal', flat=True).distinct().order_by('cor_principal')
    marcas_disponiveis = Marca.objects.all().order_by('nome')
    times_disponiveis = Time.objects.all().order_by('nome')
    temporadas_disponiveis = Camiseta.objects.values_list('temporada', flat=True).distinct()
    tipos_produto_disponiveis = TipoProduto.objects.all()
    ligas_disponiveis = Liga.objects.all().order_by('nome')

    # Renderizar o template
    return render(request, 'home.html', {
        'camisetas': camisetas,
        'cores_disponiveis': cores_disponiveis,
        'marcas_disponiveis': marcas_disponiveis,
        'times_disponiveis': times_disponiveis,
        'temporadas_disponiveis': temporadas_disponiveis,
        'tipos_produto_disponiveis': tipos_produto_disponiveis,
        'ligas_disponiveis': ligas_disponiveis,
        'filtros_aplicados': filtros_aplicados,  # Enviar os filtros ativos
    })


def sobre(request):
    return render(request, 'sobre.html')  # Carrega a página 'sobre.html'

def quem_somos(request):
    return render(request, 'quem_somos.html')  # Carrega a página 'quem_somos.html'

# Função para exportar os dados de estoque
def exportar_estoque_csv(modeladmin, request, queryset):
     # Cria o arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Estoque"

    # Cabeçalhos do arquivo Excel
    ws.append(['Time', 'Cor', 'Marca', 'Patrocinador', 'Tamanho', 'Estoque_Atual', 'Estoque_Minimo', 'Fornecedor', 'Categoria', 'Preco_de_Venda'])

    # Define a fonte vermelha e o preenchimento para estoque abaixo do mínimo
    fonte_vermelha = Font(color="FF0000")

    # Itera pelas camisetas selecionadas no queryset
    for camiseta in queryset:
        # Itera sobre os tamanhos associados a cada camiseta
        for tamanho_instance in camiseta.tamanhos.all():
            row = [
                camiseta.time.nome,
                camiseta.cor_principal,
                camiseta.marca.nome,
                camiseta.patrocinador,
                tamanho_instance.tamanho,
                tamanho_instance.quantidade_em_estoque,
                tamanho_instance.estoque_minimo,
                camiseta.fornecedor.nome if camiseta.fornecedor else 'Não definido',
                camiseta.tipo_produto.nome,
                camiseta.preco_custo
            ]
            # Adiciona a linha ao arquivo Excel
            ws.append(row)

            # Verifica se o estoque está abaixo do mínimo
            if tamanho_instance.estoque_baixo():
                # Aplica a formatação na linha correspondente
                for cell in ws[ws.max_row]:
                    cell.font = fonte_vermelha

    # Cria a resposta HTTP para baixar o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_estoque.xlsx'

    # Salva o arquivo Excel na resposta
    wb.save(response)
    return response
