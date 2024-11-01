from django.contrib import admin
import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from .models import Camiseta, CamisetaTamanho
from django.utils.html import format_html
from django.urls import reverse

class CamisetaTamanhoInline(admin.TabularInline):
    model = CamisetaTamanho
    extra = 0  # Quantas linhas extras mostrar

    # Mostrar no Django que o estoque esta abaixo de 5
    # Ainda nao esta funcionando
    def estoque_baixo_alerta(self, obj):
        if obj.estoque_baixo():
            return '⚠️ Baixo Estoque'
        return 'OK'
    estoque_baixo_alerta.short_description = 'Alerta de Estoque'


# Função para exportar os dados de estoque
def exportar_estoque_csv(modeladmin, request, queryset):
     # Cria o arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Relatório de Estoque"

    # Cabeçalhos do arquivo Excel
    ws.append(['Time', 'Cor', 'Marca', 'Patrocinador', 'Tamanho', 'Estoque_Atual', 'Estoque_Minimo', 'Fornecedor', 'Categoria', 'Preco_de_Venda'])

    # Define a fonte vermelha e o preenchimento para estoque abaixo do mínimo
    fonte_vermelha = Font(color="FF0000")

    # Itera pelas camisetas selecionadas no queryset
    for camiseta in queryset:
        # Itera sobre os tamanhos associados a cada camiseta
        for tamanho_instance in camiseta.tamanhos.all():
            row = [
                camiseta.time,
                camiseta.cor_principal,
                camiseta.marca,
                camiseta.patrocinador,
                tamanho_instance.tamanho,
                tamanho_instance.quantidade_em_estoque,
                tamanho_instance.estoque_minimo,
                camiseta.fornecedor.nome if camiseta.fornecedor else 'Não definido',
                camiseta.tipo_produto.nome,
                camiseta.preco_custo
            ]
            # Adiciona a linha ao arquivo Excel
            ws.append(row)

            # Verifica se o estoque está abaixo do mínimo
            if tamanho_instance.estoque_baixo():
                # Aplica a formatação na linha correspondente
                for cell in ws[ws.max_row]:
                    cell.font = fonte_vermelha

    # Cria a resposta HTTP para baixar o arquivo Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=relatorio_estoque.xlsx'

    # Salva o arquivo Excel na resposta
    wb.save(response)
    return response
# Nome para a ação no admin
exportar_estoque_csv.short_description = "Exportar Estoque para CSV"


class CamisetasAdmin(admin.ModelAdmin):
    list_display = ("time", "temporada", "estilo", "estoque_total", "valor_final", "link_grafico_estoque")
    list_filter = ("time", "temporada")
    inlines = [CamisetaTamanhoInline]
    actions = [exportar_estoque_csv]

    def link_grafico_estoque(self, obj):
        url = reverse('grafico_estoque', args=[obj.id])
        return format_html('<a href="{}">Gráfico Estoque</a>', url)

    link_grafico_estoque.short_description = 'Gráfico de Estoque'
    
# Register your models here.
admin.site.register(Camiseta, CamisetasAdmin)
